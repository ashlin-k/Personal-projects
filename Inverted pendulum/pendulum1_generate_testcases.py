from openpyxl import Workbook

def impulseXYZ(ws, xyzindex):

    xdd = 0.0
    ydd = 0.0
    zdd = 0.0

    for i in range (0, 50):

        if (i == 5):
            if (xyzindex == 0):
                xdd = 0.0
                ydd = 0.0
                zdd = 0.0
                omega4 = 1908.925
            elif (xyzindex == 1):
                xdd = 2.0
            elif (xyzindex == 2):
                ydd = 2.0
                omega4 = 5000.0
            elif (xyzindex == 3):
                zdd = 2.0
        else:
            xdd = 0.0
            ydd = 0.0
            zdd = 0.0

        ws.append([i+1, xdd, ydd, zdd])


def main():

    wb = Workbook()
    ws0 = wb.active
    ws0.append(['Time', 'xdd', 'ydd', 'zdd'])
    ws0.title = 'NoImpulse'
    impulseXYZ(ws0, 0)

    ws5 = wb.create_sheet()
    ws5.append(['Time', 'xdd', 'ydd', 'zdd'])
    ws5.title = 'ImpulseXdd'
    impulseXYZ(ws5, 1)

    ws1 = wb.create_sheet()
    ws1.append(['Time', 'xdd', 'ydd', 'zdd'])
    ws1.title = 'ImpulseYdd'
    impulseXYZ(ws1, 2)

    ws2 = wb.create_sheet()
    ws2.append(['Time', 'xdd', 'ydd', 'zdd'])
    ws2.title = 'ImpulseZdd'
    impulseXYZ(ws2, 3)

    wb.save("pendulum1_testcases.xlsx")


main()

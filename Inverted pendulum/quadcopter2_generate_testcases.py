from openpyxl import Workbook

def impulseOmega(ws, omegaindex):

    omega1 = 1908.925
    omega2= 1908.925
    omega3= 1908.925
    omega4 = 1908.925

    for i in range (0, 50):

        if (i == 5):
            if (omegaindex == 0):
                omega1 = 1908.925
                omega2= 1908.925
                omega3= 1908.925
                omega4 = 1908.925
            elif (omegaindex == 1):
                omega1 = 5000.0
                omega3 = 5000.0
            elif (omegaindex == 2):
                omega2 = 5000.0
                omega4 = 5000.0
            elif (omegaindex == 3):
                omega1 = 5500.0
                omega3 = 5000.0
            elif (omegaindex == 4):
                omega2 = 5500.0
                omega4 = 5000.0
            else:
                omega1 = 5000.0
                omega2 = 5000.0
                omega3 = 5000.0
                omega4 = 5000.0

        else:
            omega1 = 1908.925
            omega2= 1908.925
            omega3= 1908.925
            omega4 = 1908.925

        ws.append([i+1, omega1, omega2, omega3, omega4])


def main():

    wb = Workbook()
    ws0 = wb.active
    ws0.append(['Time', 'Omgea1', 'Omega2', 'Omega3', 'Omega4'])
    ws0.title = 'NoImpulse'
    impulseOmega(ws0, 0)

    ws5 = wb.create_sheet()
    ws5.append(['Time', 'Omgea1', 'Omega2', 'Omega3', 'Omega4'])
    ws5.title = 'ImpulseOmega13'
    impulseOmega(ws5, 1)

    ws1 = wb.create_sheet()
    ws1.append(['Time', 'Omgea1', 'Omega2', 'Omega3', 'Omega4'])
    ws1.title = 'ImpulseOmega24'
    impulseOmega(ws1, 2)

    ws2 = wb.create_sheet()
    ws2.append(['Time', 'Omgea1', 'Omega2', 'Omega3', 'Omega4'])
    ws2.title = 'ImpulsePitchY'
    impulseOmega(ws2, 3)

    ws3 = wb.create_sheet()
    ws3.append(['Time', 'Omgea1', 'Omega2', 'Omega3', 'Omega4'])
    ws3.title = 'ImpulseRollX'
    impulseOmega(ws3, 4)

    ws4 = wb.create_sheet()
    ws4.append(['Time', 'Omgea1', 'Omega2', 'Omega3', 'Omega4'])
    ws4.title = 'ImpulseAllOmega'
    impulseOmega(ws4, 5)

    wb.save("quadcopter2_testcases.xlsx")


main()

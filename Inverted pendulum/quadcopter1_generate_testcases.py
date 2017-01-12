from openpyxl import Workbook

def noStimulus(ws):

    thrust = 9.81
    omegax = 0.0
    omegay= 0.0
    omegaz= 0.0

    for i in range (0, 100):

        ws.append([i+1, omegax, omegay, omegaz, thrust])
        

def impulseOmega(ws, omegadir):

    thrust = 9.81
    omegax = 0.0
    omegay= 0.0
    omegaz= 0.0

    for i in range (0, 100):

        if (i == 10):
            if (omegadir == 1):
                omegax = 2.0
            elif (omegadir == 2):
                omegay = 2.0
            else:
                omegaz = 2.0
        else:
            omegax = 0.0
            omegay = 0.0
            omegaz = 0.0

        ws.append([i+1, omegax, omegay, omegaz, thrust])


def impulseThrust(ws):

    thrust = 9.81
    omegax = 0.0
    omegay= 0.0
    omegaz= 0.0

    for i in range (0, 100):

        if (i == 10):
            thrust = 50.0
        else:
            thrust = 9.81

        ws.append([i+1, omegax, omegay, omegaz, thrust])

    

def main():

    wb = Workbook()
    ws0 = wb.active
    ws0.append(['Time', 'Omgea_x', 'Omega_y', 'Omega_z', 'Thrust'])
    ws0.title = 'Impulse omega'
    noStimulus(ws0)

    ws1 = wb.create_sheet()
    ws1.append(['Time', 'Omgea_x', 'Omega_y', 'Omega_z', 'Thrust'])
    ws1.title = 'Impulse omeag X'
    impulseOmega(ws1, 1)

    ws2 = wb.create_sheet()
    ws2.append(['Time', 'Omgea_x', 'Omega_y', 'Omega_z', 'Thrust'])
    ws2.title = 'Impulse omega Y'
    impulseOmega(ws2, 2)

    ws3 = wb.create_sheet()
    ws3.append(['Time', 'Omgea_x', 'Omega_y', 'Omega_z', 'Thrust'])
    ws3.title = 'Impulse omega Z'
    impulseOmega(ws3, 3)

    ws4 = wb.create_sheet()
    ws4.append(['Time', 'Omgea_x', 'Omega_y', 'Omega_z', 'Thrust'])
    ws4.title = 'Impulse thrust'
    impulseThrust(ws4)

    wb.save("quadcopter1_testcases.xlsx")


main()
